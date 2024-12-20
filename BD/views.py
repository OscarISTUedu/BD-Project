import json
from datetime import datetime
from types import NoneType

import openpyxl
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.db.models import ManyToOneRel, OuterRef, Subquery
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect

from django.apps import apps
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .authorisation import group_required
from .models import Patient, Doctor, Neighborhood, Diagnosis, Visit, Ticket
from .validators import validate_status, validate_patient_id, validate_doctor_id, validate_empty

def transliterate(text):
    translit_dict = {
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo', 'Ж': 'Zh', 'З': 'Z',
        'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R',
        'С': 'S', 'Т': 'T', 'У': 'U', 'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
        'Ы': 'Y', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',

        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo', 'ж': 'zh', 'з': 'z',
        'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
        'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ы': 'y', 'э': 'e', 'ю': 'yu', 'я': 'ya'
    }
    return ''.join([translit_dict.get(char, char) for char in text])

@permission_required('BD.view_patient')
def patient_list(request):
    patients = Patient.objects.all().order_by('id').values()
    fields = Patient._meta.get_fields()
    fields = fields[1:]

    return render(request, 'model_list.html',
                  {'models': patients,'h1':Patient._meta.verbose_name_plural,'fields':fields,"fields_len":len(fields)})

@permission_required('BD.view_doctor')
def doctor_list(request):
    doctors = Doctor.objects.all().order_by('id').values()
    fields = Doctor._meta.get_fields()
    fields = fields[1:]
    return render(request, 'model_list.html',
                  {'models': doctors, 'h1': Doctor._meta.verbose_name_plural, 'fields': fields,"fields_len":len(fields)})

@permission_required('BD.view_neighborhood')
def neighborhood_list(request):
    neighborhoods = Neighborhood.objects.all().order_by('id').values()
    fields = Neighborhood._meta.get_fields()
    fields = fields[1:]
    return render(request, 'model_list.html',
                  {'models': neighborhoods, 'h1': Neighborhood._meta.verbose_name_plural, 'fields': fields,"fields_len":len(fields)})

@permission_required('BD.view_diagnosis')
def diagnoses_list(request):
    diagnoses = Diagnosis.objects.all().order_by('id').values()
    fields = Diagnosis._meta.get_fields()
    fields = fields[1:]
    return render(request, 'model_list.html',
                  {'models': diagnoses, 'h1': Diagnosis._meta.verbose_name_plural, 'fields': fields,"fields_len":len(fields)})

@permission_required('BD.view_visit')
def visit_list(request):
    visits = Visit.objects.all().order_by('id').values()
    fields = Visit._meta.get_fields()
    fields = fields[1:]
    return render(request, 'model_list.html',
                  {'models': visits, 'h1': Visit._meta.verbose_name_plural, 'fields': fields,"fields_len":len(fields)})

@permission_required('BD.view_ticket')
def ticket_list(request):
    tickets = Ticket.objects.all().order_by('id').values()
    Patient.objects.filter()
    fields = Ticket._meta.get_fields()
    for ticket in tickets:
        if ticket.get('diagnosis_id'):
            diagnosis = Diagnosis.objects.filter(id=ticket.get('diagnosis_id')).first().diagnosis
        else:
            diagnosis = None
        visit = Visit.objects.filter(id=ticket.get('visit_id')).first().visit
        patient = Patient.objects.filter(id=ticket.get('patient_id')).first().surname
        doctor = Doctor.objects.filter(id=ticket.get('doctor_id')).first().surname
        ticket['diagnosis_id'] = diagnosis
        ticket['visit_id'] =  visit
        ticket['patient_id'] = patient +","+"№"+str(ticket['patient_id'])
        ticket['doctor_id'] = doctor +","+ "№"+str(ticket['doctor_id'])
    return render(request, 'model_list.html',
                  {'models': tickets, 'h1': Ticket._meta.verbose_name_plural, 'fields': fields,"fields_len":len(fields)})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('main')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def main_view(request):
    return render(request, 'main.html')

@login_required #Изменение текстовых полей
def change_view(request):
    data = json.loads(request.body)
    new_data = data.get('new_data')
    last_data = data.get('last_data')
    if new_data==last_data:
        return HttpResponse(status=304)
    verbose_name_plural = data.get('table_verbose_name_plural')
    verbose_name_field = data.get('field_name')
    row_id = data.get('id')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == verbose_name_plural:
            cur_model = model
            break
    fields = cur_model._meta.get_fields()
    for field in fields:
        if not isinstance(field, ManyToOneRel):
            if field.verbose_name == verbose_name_field:
                cur_field_name = field.name
                break
    field = cur_model._meta.get_field(cur_field_name)
    if new_data=="-":
        if field.null and field.blank:
            new_data = None
        else:
            return JsonResponse({"response": "Не верный формат"}, status=500)
    cur_obj = cur_model.objects.filter(id=row_id).first()
    if cur_obj:#если сущ-ет, то меняем, если нет, то
        try:
            setattr(cur_obj, cur_field_name, new_data)
        except Exception as e:
            return JsonResponse({"response": f"Значение {new_data} не найдено в дочерней таблице,возмонжо проблема в некоректной форме"}, status=500)
        try:
            cur_obj.full_clean()
            cur_obj.save()
        except ValidationError as e:
            return JsonResponse({"response":e.messages[0]},status=500)
    else:
        try:
            field.clean(new_data, None)
        except Exception as e:
            return JsonResponse({"response": "Не верный формат"},status=500)
    return JsonResponse({"key":cur_field_name,"value":new_data},status=200)

@login_required
def add_empty_row(request):
    data = json.loads(request.body)
    verbose_name_plural = data.pop('table_verbose_name_plural')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == verbose_name_plural:
            cur_model = model
            break
    last_obj = cur_model.objects.last().id
    return JsonResponse({"id":last_obj+1},status=200)

@login_required
def validate_field (request):#валидация внешних ключей + статуса, данной записи нет в бд, но есть в new_row
    data = json.loads(request.body)
    new_data = data.get('new_data')
    last_data = data.get('last_data')
    type = data.get('type')
    if new_data == last_data and ((type == "text&id" and new_data != "-") or (type == "id" and new_data != "-")):
        return HttpResponse(status=304)
    model_name = data.get('model_name')
    new_row = data.get('new_row')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == model_name:
            cur_model = model
            break
    return validate_empty(data, cur_model, new_row)

@login_required
def get_fields_by_name(request):#все значения для выпадающего списка
    data = json.loads(request.body)
    field_name = data.pop('field_name')
    model_name = data.pop('model_name')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == model_name:
            cur_model = model
            break
    relatedModel = cur_model._meta.get_field(field_name).related_model
    field_cut = field_name.split('_')
    if field_name in ["visit_id","diagnosis_id"]:#выводятся - названия, значения - id
        values = list(relatedModel.objects.values_list(field_cut[1],field_cut[0]))
        values = [item for item in values if item[0] is not None]#убрали None
        values.append((-1,"-")) if cur_model._meta.get_field(field_name).null and cur_model._meta.get_field(field_name).blank else None
        values = sorted(values, key=lambda x: x[0])
        return JsonResponse({"values": values,"type":"text&id"}, status=200)
    elif field_name in ["neighborhood_id"]:
        field_cut[0]+="_street"
        values = list(relatedModel.objects.values_list(field_cut[1],flat=True))
        values.sort()
        return JsonResponse({"values": values,"type":"id"}, status=200)
    elif field_name in ["category"]:
        values = cur_model.objects.values_list(field_name, flat=True).distinct()
        values = [item for item in values if item is not None]
        values.append("-") if cur_model._meta.get_field(field_name).null and cur_model._meta.get_field(field_name).blank else None
    elif field_name in ["surname"] and cur_model._meta.verbose_name_plural=="Врачи":
        values = Doctor.objects.values_list("id", "surname")
        values = [item for item in values if item[0] is not None]
        values = [(value1, value2 + ",№" + str(value1)) for value1, value2 in values]
        values = sorted(values, key=lambda x: x[0])
        return JsonResponse({"values": values, "type": "text&id"}, status=200)
    elif field_name in ["diagnosis"]:
        values = list(Diagnosis.objects.values_list("id", "diagnosis"))
        values = [item for item in values if item[0] is not None]  # убрали None
        values.append((-1, "-")) if Diagnosis._meta.get_field(field_name).null and Diagnosis._meta.get_field(field_name).blank else None
        values = sorted(values, key=lambda x: x[0])
        return JsonResponse({"values": values,"type":"text&id"}, status=200)
    else:#статус, улица(пациента)
        values = cur_model.objects.values_list(field_name,flat=True).distinct()
    if field_name in ["doctor_id","patient_id"]:
        values = relatedModel.objects.values_list("id","surname")
        values = [item for item in values if item[0] is not None]
        values = [(value1,value2+",№"+str(value1)) for value1,value2 in values]
        values = sorted(values, key=lambda x: x[0])
        return JsonResponse({"values": values, "type": "text&id"}, status=200)
    try:
        values = [int(val) for val in values]
    except Exception:
        values = [val for val in values]
    values.sort()
    return JsonResponse({"values": values,"type":"id"}, status=200)

@login_required
def change_by_list(request):#изменение (не)существующей записи в бд, её валидация
    data = json.loads(request.body)
    new_data = data.get('new_data')
    last_data = data.get('last_data')
    type = data.get('type')
    if new_data == last_data and ((type == "text&id" and new_data!="-") or (type == "id" and new_data!="-")):
        return HttpResponse(status=304)
    model_name = data.get('model_name')
    field_name = data.get('field_name')
    row_id = data.get('id')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == model_name:
            cur_model = model
            break
    cur_obj = cur_model.objects.filter(id=row_id).first()
    if type == "text&id":
        field_id = data.get('field_id')#id в родительской таблице
        field_id = None if new_data=="-" else field_id
        relatedModel = cur_model._meta.get_field(field_name).related_model
        validated_patient_id = validate_patient_id(field_name,relatedModel,field_id,cur_obj,{})
        if not isinstance(validated_patient_id,NoneType):
            return validated_patient_id
        validated_doctor_id = validate_doctor_id(field_name,relatedModel,field_id,cur_obj,{})
        if not isinstance(validated_doctor_id,NoneType):
            return validated_doctor_id
        try:
            setattr(cur_obj, field_name, field_id)
        except Exception as e:
            return JsonResponse(
                {"response": f"Значение {field_id} не найдено в родительской таблице,возможно проблема в некоректной форме"},
                status=500)
    else:
        new_data = None if new_data == "-" else new_data
        validated_status = validate_status(field_name,cur_obj,new_data,{})#Проверка валидности поля status
        if not isinstance(validated_status,NoneType):
            return validated_status
        try:
            setattr(cur_obj, field_name, new_data)
        except Exception as e:
            return JsonResponse(
                {"response": f"Значение {new_data} не найдено в родительской таблице,возможно проблема в некоректной форме"},
                status=500)
    try:
        cur_obj.full_clean()
        cur_obj.save()
    except ValidationError as e:
        return JsonResponse({"response": e.messages[0]}, status=500)
    return JsonResponse({},status=200)

@login_required
def row_add(request):
    data = json.loads(request.body)
    model_name = data.pop('model_name')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == model_name:
            cur_model = model
            break
    try:
        cur_obj = cur_model.objects.create(**data)
        cur_obj.full_clean()
        cur_obj.save()
    except Exception as e:
        return JsonResponse({"Ошибка"}, status=500)
    return JsonResponse({},status=200)

@login_required
def row_delete(request):
    data = json.loads(request.body)
    row_id = data.get('id')
    model_name = data.get('model_name')
    for model in apps.get_models():
        if model._meta.verbose_name_plural == model_name:
            cur_model = model
            break
    try:
        cur_model.objects.filter(id=row_id).first().delete()
    except Exception as e:
        return JsonResponse({"Ошибка удаления элемента"},status=500)
    return JsonResponse({},status=200)


def get_str_neigh_dict(request):
    neighborhoods = list(Neighborhood.objects.all().values())
    neighborhoods = dict([(str(id["id"]), id["neighborhood_street"]) for id in neighborhoods])
    return JsonResponse (neighborhoods,status=200)
'''
Выходные документы
'''

@group_required(['Пациенты','Администрация'])
def doc_neigh_doc(request):#Список участков и участковых врачей
    wb = openpyxl.Workbook()
    sheet  = wb.active
    sheet.title = "Участки и врачи"
    headers = ["Номер участка", "Имя доктора", "Фамилия доктора","Специальность доктора"]
    list_title = "Список участков и участковых врачей"
    merge_range = f'A1:{get_column_letter(len(headers))}1'
    sheet.merge_cells(merge_range)
    sheet['A1'] = list_title
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):#Выравнивание
        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    neighborhoods = Neighborhood.objects.prefetch_related('doctor').order_by('id')
    for neighborhood in neighborhoods:
        for doctor in neighborhood.doctor.all():
            sheet.append([neighborhood.id, doctor.name, doctor.surname, doctor.speciality])
    for column in sheet.columns:#увеличение ширины столбцов
        max_length = 0
        column_letter = column[1].column_letter  # Получаем букву столбца
        for cell in column:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value and cell.value!=list_title:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2  # Добавляем небольшой отступ
    last_row = sheet.max_row
    start_cell = f'A{last_row + 1}'  # Первая ячейка в строке
    end_column = get_column_letter(len(headers))  # Получаем букву i-го столбца
    end_cell = f'{end_column}{last_row + 1}'  # Последняя ячейка в строке
    sheet.merge_cells(f'{start_cell}:{end_cell}')
    sheet[f'{start_cell}'] = 'Дата: ' + str(datetime.today().date())
    sheet[f'{start_cell}'].alignment = Alignment(horizontal="center", vertical="center")
    # Создание ответа с Excel-файлом
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="doc_neigh_doc.xlsx"'
    # Сохранение книги в ответ
    wb.save(response)
    return response

@group_required(['Регистратура'])
def ticket_print(request):#Вывод талонов
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Таблица талонов"
    headers = ["Номер", "Дата и время приёма", "Врач,номер", "Пациент,номер","Цель посещения","Диагноз","Статус"]
    list_title = "Список талонов"
    merge_range = f'A1:{get_column_letter(len(headers))}1'
    sheet.merge_cells(merge_range)
    sheet['A1'] = list_title
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):#Выравнивание
        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    tickets = Ticket.objects.all().order_by('id')
    for ticket in tickets:
        diagnosis = "-" if ticket.diagnosis is None else ticket.diagnosis.diagnosis
        sheet.append([ticket.id, ticket.date_n_time,str(ticket.doctor.surname)+",№"+str(ticket.doctor.id) ,str(ticket.patient.surname)+",№"+str(ticket.patient.id),ticket.visit.visit,diagnosis,ticket.status])
    for column in sheet.columns:
        max_length = 0
        column_letter = column[1].column_letter
        for cell in column:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value and cell.value!=list_title:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2
    last_row = sheet.max_row
    start_cell = f'A{last_row + 1}'  # Первая ячейка в строке
    end_column = get_column_letter(len(headers))  # Получаем букву i-го столбца
    end_cell = f'{end_column}{last_row + 1}'  # Последняя ячейка в строке
    sheet.merge_cells(f'{start_cell}:{end_cell}')
    sheet[f'{start_cell}'] = 'Дата: ' + str(datetime.today().date())
    sheet[f'{start_cell}'].alignment = Alignment(horizontal="center", vertical="center")
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ticket_print.xlsx"'
    wb.save(response)
    return response

@group_required(['Администрация'])
def patient_diagnosis(request):#Вывод списка пациентов с определённым диагнозом
    data = json.loads(request.body)
    diagnosis_id = data.get('id')
    diagnosis_name_rus = Diagnosis.objects.filter(id=diagnosis_id).first().diagnosis
    diagnosis_name = transliterate(diagnosis_name_rus)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Пациенты"
    headers = ["Номер","Имя", "Фамилия", "Отчество", "Пол", "Дата рождения"]
    list_title = f"Список пациентов с диагнозом {diagnosis_name_rus}"
    merge_range = f'A1:{get_column_letter(len(headers))}1'
    sheet.merge_cells(merge_range)
    sheet['A1'] = list_title
    sheet.append(headers)
    for col_num, header in enumerate(headers, start=1):#Выравнивание
        cell = sheet.cell(row=1, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    tickets = Ticket.objects.filter(diagnosis = diagnosis_id)
    for ticket in tickets:
        third_name = "-" if ticket.patient.third_name is None else ticket.patient.third_name
        sheet.append([ticket.patient.id, ticket.patient.name, ticket.patient.surname, third_name, ticket.patient.sex, ticket.patient.date_of_birth])
    for column in sheet.columns:
        max_length = 0
        column_letter = column[1].column_letter
        for cell in column:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value and cell.value!=list_title:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2
    last_row = sheet.max_row
    start_cell = f'A{last_row + 1}'  # Первая ячейка в строке
    end_column = get_column_letter(len(headers))  # Получаем букву i-го столбца
    end_cell = f'{end_column}{last_row + 1}'  # Последняя ячейка в строке
    sheet.merge_cells(f'{start_cell}:{end_cell}')
    sheet[f'{start_cell}'] = 'Дата: ' + str(datetime.today().date())
    sheet[f'{start_cell}'].alignment = Alignment(horizontal="center", vertical="center")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename=\"patient_{diagnosis_name}.xlsx\""
    wb.save(response)
    return response

@group_required(['Администрация'])
def patient_doctor(request):#Вых.док - Вывод списка пациентов побывавших на приеме у определенного врача за определенный период
    data = json.loads(request.body)
    data = data['DataTime']
    if len(data)==3:
        doctor_id = data['id']
        doctor_surname_rus = Doctor.objects.filter(id=doctor_id).first().surname
        doctor_surname = transliterate(doctor_surname_rus)
        data_time_start = data['StartDataTimeFrame']
        data_time_end = data['EndDataTimeFrame']
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Пациенты"
        headers = ["Номер", "Имя", "Фамилия", "Отчество", "Пол", "Дата рождения","Диагноз","Дата посещения"]
        list_title = f"Список пациентов побывавших на приеме у {doctor_surname_rus} за {data_time_start} - {data_time_end}"
        merge_range = f'A1:{get_column_letter(len(headers))}1'
        sheet.merge_cells(merge_range)
        sheet['A1'] = list_title
        sheet.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        tickets = Ticket.objects.filter(doctor=doctor_id,date_n_time__range=[data_time_start,data_time_end],diagnosis__isnull=False)
        for ticket in tickets:
            third_name = "-" if ticket.patient.third_name is None else ticket.patient.third_name
            sheet.append([ticket.patient.id, ticket.patient.name, ticket.patient.surname, third_name, ticket.patient.sex, ticket.patient.date_of_birth, ticket.diagnosis.diagnosis, ticket.date_n_time])
        for column in sheet.columns:
            max_length = 0
            column_letter = column[1].column_letter
            for cell in column:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value and cell.value!=list_title:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2
        last_row = sheet.max_row
        start_cell = f'A{last_row + 1}'  # Первая ячейка в строке
        end_column = get_column_letter(len(headers))  # Получаем букву i-го столбца
        end_cell = f'{end_column}{last_row + 1}'  # Последняя ячейка в строке
        sheet.merge_cells(f'{start_cell}:{end_cell}')
        sheet[f'{start_cell}'] = 'Дата: ' + str(datetime.today().date())
        sheet[f'{start_cell}'].alignment = Alignment(horizontal="center", vertical="center")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=\"patients_{doctor_surname}.xlsx\""
        wb.save(response)
        return response
    return JsonResponse ({},status=200)